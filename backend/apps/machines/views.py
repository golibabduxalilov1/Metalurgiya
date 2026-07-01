"""
Machines App Views
"""
import io
import os
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from django.utils import timezone
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.permissions import IsAdmin, IsAdminOrMaster, IsAdminOrMasterOrOwner
from .models import (
    Machine, MachineType, MachineStatus,
    MachineStatusHistory, MachineAttachment, MachineAssignment,
    MaintenanceSchedule, RepairTask, TaskSparePart, MaintenanceHistory
)
from .serializers import (
    MachineListSerializer, MachineDetailSerializer, MachineCreateUpdateSerializer,
    MachineTypeSerializer, MachineStatusSerializer, MachineStatusHistorySerializer,
    MachineAttachmentSerializer, MachineAttachmentCreateSerializer,
    MachineAssignmentSerializer, ChangeStatusSerializer,
    MaintenanceScheduleSerializer, MaintenanceScheduleWriteSerializer, MaintenanceCompleteSerializer,
    RepairTaskSerializer, TaskSparePartSerializer, MaintenanceHistorySerializer,
)
from .filters import MachineFilter
from apps.audit.signals import log_action


@extend_schema(tags=['machine-types'])
class MachineTypeViewSet(viewsets.ModelViewSet):
    """Управление типами станков"""
    queryset = MachineType.objects.all()
    serializer_class = MachineTypeSerializer
    permission_classes = [IsAdminOrMaster]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return []
        return [IsAdmin()]


@extend_schema(tags=['machine-statuses'])
class MachineStatusViewSet(viewsets.ModelViewSet):
    """Управление статусами станков"""
    queryset = MachineStatus.objects.all()
    serializer_class = MachineStatusSerializer
    search_fields = ['name']
    ordering_fields = ['sort_order', 'name']

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return []
        return [IsAdmin()]


@extend_schema(tags=['machines'])
@extend_schema_view(
    list=extend_schema(
        summary='Список станков',
        parameters=[
            OpenApiParameter('status', description='ID статуса'),
            OpenApiParameter('workshop', description='ID цеха'),
            OpenApiParameter('machine_type', description='ID типа'),
            OpenApiParameter('operator', description='ID оператора'),
            OpenApiParameter('search', description='Поиск по тексту'),
            OpenApiParameter('include_deleted', description='Включать удалённые', type=bool),
        ]
    ),
    retrieve=extend_schema(summary='Карточка станка'),
    create=extend_schema(summary='Добавить станок'),
    update=extend_schema(summary='Обновить станок'),
    partial_update=extend_schema(summary='Частично обновить станок'),
)
class MachineViewSet(viewsets.ModelViewSet):
    """
    Реестр станков. Поддерживает:
    - CRUD операции
    - Поиск, фильтрация, сортировка
    - Смена статуса с историей
    - Прикрепление файлов
    - Закрепление операторов
    - Импорт/экспорт Excel
    - Мягкое удаление и восстановление
    """
    queryset = Machine.objects.select_related(
        'machine_type', 'current_status', 'workshop', 'section',
        'assigned_operator', 'created_by'
    ).prefetch_related('attachments', 'status_history', 'assignments')
    filterset_class = MachineFilter
    search_fields = ['name', 'inventory_number', 'model', 'manufacturer', 'description']
    ordering_fields = ['name', 'inventory_number', 'created_at', 'updated_at']
    ordering = ['inventory_number']

    def get_permissions(self):
        # Read-only actions — any authenticated user
        if self.action in [
            'list', 'retrieve', 'period_stats', 'export_excel', 'import_template',
            'status_history', 'maintenance', 'maintenance_tasks',
            'maintenance_task_detail', 'maintenance_task_spare_parts',
            'maintenance_task_spare_part_delete', 'maintenance_history',
        ]:
            return [IsAdminOrMasterOrOwner()]
        if self.action in ['create', 'update', 'partial_update', 'assign_operator',
                          'change_status', 'upload_attachment']:
            return [IsAdminOrMaster()]
        if self.action in ['destroy', 'restore', 'import_excel']:
            return [IsAdmin()]
        return [IsAdminOrMaster()]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return MachineDetailSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return MachineCreateUpdateSerializer
        return MachineListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        if not include_deleted:
            qs = qs.filter(deleted_at__isnull=True)
        elif not self.request.user.is_admin:
            qs = qs.filter(deleted_at__isnull=True)
        return qs

    def perform_create(self, serializer):
        machine = serializer.save(created_by=self.request.user)
        log_action(self.request.user, 'create', machine, None, machine.__dict__)

    def perform_update(self, serializer):
        old_data = {f.name: getattr(serializer.instance, f.name) for f in serializer.instance._meta.fields}
        machine = serializer.save()
        new_data = {f.name: getattr(machine, f.name) for f in machine._meta.fields}
        log_action(self.request.user, 'update', machine, old_data, new_data)

    @extend_schema(summary='Мягкое удаление станка')
    def destroy(self, request, *args, **kwargs):
        machine = self.get_object()
        try:
            machine.maintenance_schedule.delete()
        except MaintenanceSchedule.DoesNotExist:
            pass
        MaintenanceHistory.objects.filter(machine=machine).delete()
        machine.deleted_at = timezone.now()
        machine.deleted_by = request.user
        machine.save()
        log_action(request.user, 'delete', machine, None, None)
        return Response({'detail': 'Станок помечен как удалённый'}, status=status.HTTP_200_OK)

    @extend_schema(summary='Восстановить удалённый станок')
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        machine = Machine.objects.get(pk=pk)
        machine.deleted_at = None
        machine.deleted_by = None
        machine.save()
        log_action(request.user, 'restore', machine, None, None)
        return Response({'detail': 'Станок восстановлен'})

    @extend_schema(summary='Сменить статус станка', request=ChangeStatusSerializer)
    @action(detail=True, methods=['post'], url_path='change-status')
    def change_status(self, request, pk=None):
        machine = self.get_object()
        serializer = ChangeStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data['status']
        comment = serializer.validated_data.get('comment', '')
        old_status = machine.current_status

        # Create history record
        MachineStatusHistory.objects.create(
            machine=machine,
            previous_status=old_status,
            new_status=new_status,
            changed_by=request.user,
            comment=comment
        )

        machine.current_status = new_status
        machine.save(update_fields=['current_status', 'updated_at'])
        log_action(request.user, 'status_change', machine,
                   {'status': str(old_status)}, {'status': str(new_status)})

        return Response(MachineDetailSerializer(machine, context={'request': request}).data)

    @extend_schema(summary='История статусов станка')
    @action(detail=True, methods=['get'], url_path='status-history')
    def status_history(self, request, pk=None):
        machine = self.get_object()
        history = machine.status_history.all()
        serializer = MachineStatusHistorySerializer(history, many=True)
        return Response(serializer.data)

    @extend_schema(summary='Статистика работы станков за период')
    @action(detail=False, methods=['get'], url_path='period-stats')
    def period_stats(self, request):
        import datetime as dt
        from django.utils.dateparse import parse_date
        from django.utils import timezone as tz_utils

        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')

        today = tz_utils.now().date()
        d_from = parse_date(date_from_str) if date_from_str else today - dt.timedelta(days=6)
        d_to = parse_date(date_to_str) if date_to_str else today

        if not d_from or not d_to or d_from > d_to:
            return Response({'detail': 'Неверный диапазон дат'}, status=status.HTTP_400_BAD_REQUEST)

        dt_from = dt.datetime.combine(d_from, dt.time.min, tzinfo=dt.timezone.utc)
        # Bugungi kun uchun — hozirgi vaqtgacha, o'tgan kunlar uchun — kun oxirigacha
        if d_to >= today:
            dt_to = tz_utils.now()
        else:
            dt_to = dt.datetime.combine(d_to, dt.time.max, tzinfo=dt.timezone.utc)

        machines = Machine.objects.filter(deleted_at__isnull=True).select_related('current_status')

        num_days = (d_to - d_from).days + 1
        daily_buckets = [
            {
                'date': (d_from + dt.timedelta(days=i)).strftime('%d.%m.%Y'),
                'green': 0.0, 'red': 0.0, 'yellow': 0.0,
            }
            for i in range(num_days)
        ]
        color_totals = {'green': 0.0, 'red': 0.0, 'yellow': 0.0}

        def add_segment(seg_s, seg_e, color):
            if seg_s >= seg_e or color not in color_totals:
                return
            color_totals[color] += (seg_e - seg_s).total_seconds()
            cur = seg_s
            while cur < seg_e:
                idx = (cur.date() - d_from).days
                next_day = dt.datetime.combine(
                    cur.date() + dt.timedelta(days=1), dt.time.min, tzinfo=dt.timezone.utc
                )
                chunk_end = min(seg_e, next_day)
                if 0 <= idx < num_days:
                    daily_buckets[idx][color] += (chunk_end - cur).total_seconds()
                cur = next_day

        for machine in machines:
            history = list(
                MachineStatusHistory.objects.filter(
                    machine=machine,
                    changed_at__lte=dt_to,
                ).select_related('new_status', 'previous_status').order_by('changed_at')
            )

            last_before = None
            in_period = []
            for h in history:
                if h.changed_at < dt_from:
                    last_before = h
                else:
                    in_period.append(h)

            # Stanok yaratilgan vaqtdan oldingi vaqtni sanama
            machine_start = max(dt_from, machine.created_at) if machine.created_at else dt_from
            # Stanok davr tugagandan keyin yaratilgan bo'lsa — o'tkazib yubor
            if machine_start >= dt_to:
                continue

            if last_before and last_before.new_status:
                # Davrdan OLDIN oxirgi holat o'zgarishi → o'sha holat bilan boshlash
                cur_color = last_before.new_status.color
            elif in_period and in_period[0].previous_status:
                # Davrdan oldin tarix yo'q, lekin davr ichida o'zgarish bor →
                # birinchi o'zgarishning oldingi holati = davr boshidagi haqiqiy holat
                cur_color = in_period[0].previous_status.color
            elif not in_period:
                # Hech qanday tarix yo'q → joriy holat stanok yaratilganidan beri saqlanib kelgan
                cur_color = machine.current_status.color if machine.current_status else 'gray'
            else:
                # Birinchi o'zgarishning previous_status si null → holat noma'lum, sanama
                cur_color = None

            prev = machine_start  # dt_from emas, stanok yaratilgan vaqtdan boshlash
            for hist in in_period:
                # Davrdan oldin bo'lgan o'zgarishlarni o'tkazib yuborish
                if hist.changed_at <= machine_start:
                    cur_color = hist.new_status.color if hist.new_status else 'gray'
                    continue
                add_segment(prev, hist.changed_at, cur_color)
                cur_color = hist.new_status.color if hist.new_status else 'gray'
                prev = hist.changed_at
            add_segment(prev, dt_to, cur_color)

        def to_h(s):
            return round(s / 3600, 1)

        total_s = sum(color_totals.values())

        return Response({
            'date_from': d_from.isoformat(),
            'date_to': d_to.isoformat(),
            'working_hours': to_h(color_totals['green']),
            'repair_hours': to_h(color_totals['red']),
            'idle_hours': to_h(color_totals['yellow']),
            'total_hours': to_h(total_s),
            'working_pct': round(color_totals['green'] / total_s * 100, 1) if total_s else 0,
            'repair_pct': round(color_totals['red'] / total_s * 100, 1) if total_s else 0,
            'idle_pct': round(color_totals['yellow'] / total_s * 100, 1) if total_s else 0,
            'daily': [{
                'date': b['date'],
                'working_hours': to_h(b['green']),
                'repair_hours': to_h(b['red']),
                'idle_hours': to_h(b['yellow']),
            } for b in daily_buckets],
        })

    @extend_schema(summary='Загрузить файл к станку', request=MachineAttachmentCreateSerializer)
    @action(detail=True, methods=['post'], url_path='upload',
            parser_classes=[MultiPartParser, FormParser])
    def upload_attachment(self, request, pk=None):
        machine = self.get_object()
        serializer = MachineAttachmentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = request.FILES['file']
        attachment = MachineAttachment.objects.create(
            machine=machine,
            file=file,
            original_filename=file.name,
            file_size=file.size,
            content_type=file.content_type,
            description=serializer.validated_data.get('description', ''),
            uploaded_by=request.user
        )
        return Response(
            MachineAttachmentSerializer(attachment, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )

    @extend_schema(summary='Закрепить оператора за станком')
    @action(detail=True, methods=['post'], url_path='assign-operator')
    def assign_operator(self, request, pk=None):
        from apps.employees.models import Employee
        machine = self.get_object()
        operator_id = request.data.get('operator_id')
        notes = request.data.get('notes', '')

        # Unassign current
        MachineAssignment.objects.filter(machine=machine, is_current=True).update(
            is_current=False, unassigned_at=timezone.now()
        )

        if operator_id:
            try:
                operator = Employee.objects.get(id=operator_id)
            except Employee.DoesNotExist:
                return Response({'detail': 'Оператор не найден'}, status=status.HTTP_404_NOT_FOUND)

            assignment = MachineAssignment.objects.create(
                machine=machine,
                operator=operator,
                assigned_by=request.user,
                notes=notes
            )
            machine.assigned_operator = operator
            machine.save(update_fields=['assigned_operator', 'updated_at'])
            return Response(MachineAssignmentSerializer(assignment).data, status=status.HTTP_201_CREATED)
        else:
            machine.assigned_operator = None
            machine.save(update_fields=['assigned_operator', 'updated_at'])
            return Response({'detail': 'Оператор откреплён'})

    @extend_schema(summary='Экспорт реестра в Excel')
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        machines = self.filter_queryset(self.get_queryset())
        wb = self._build_excel_export(machines)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="machines_export.xlsx"'
        return response

    def _build_excel_export(self, machines):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Реестр станков'

        headers = [
            'Инв. №', 'Наименование', 'Модель', 'Производитель',
            'Год выпуска', 'Тип станка', 'Статус', 'Цех', 'Участок',
            'Рабочее место', 'Оператор', 'Дата ввода', 'Описание'
        ]

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[1].height = 40

        for row_num, machine in enumerate(machines, 2):
            row_data = [
                machine.inventory_number,
                machine.name,
                machine.model or '',
                machine.manufacturer or '',
                machine.year_manufactured or '',
                machine.machine_type.name if machine.machine_type else '',
                machine.current_status.name if machine.current_status else '',
                machine.workshop.name if machine.workshop else '',
                machine.section.name if machine.section else '',
                machine.workplace or '',
                machine.assigned_operator.full_name if machine.assigned_operator else '',
                machine.commissioned_date.strftime('%d.%m.%Y') if machine.commissioned_date else '',
                machine.description or '',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')

        col_widths = [15, 30, 20, 20, 12, 20, 15, 20, 20, 15, 25, 15, 40]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A2'
        return wb

    @extend_schema(summary='Импорт станков из Excel')
    @action(detail=False, methods=['post'], url_path='import-excel',
            parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        if 'file' not in request.FILES:
            return Response({'detail': 'Файл не предоставлен'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        errors = []
        created = 0
        updated = 0

        default_status = MachineStatus.objects.filter(color='green', is_active=True).first()

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                inv_num = str(row[0]).strip() if row[0] else None
                if not inv_num:
                    errors.append({'row': row_num, 'error': 'Инвентарный номер обязателен'})
                    continue

                try:
                    machine, created_flag = Machine.objects.update_or_create(
                        inventory_number=inv_num,
                        defaults={
                            'name': str(row[1] or '').strip() or inv_num,
                            'model': str(row[2] or '').strip(),
                            'manufacturer': str(row[3] or '').strip(),
                            'created_by': request.user,
                        }
                    )
                    save_fields = []
                    if created_flag:
                        if default_status:
                            machine.current_status = default_status
                            save_fields.append('current_status')
                        created += 1
                    else:
                        if machine.deleted_at is not None:
                            machine.deleted_at = None
                            machine.deleted_by = None
                            save_fields += ['deleted_at', 'deleted_by']
                            if default_status and not machine.current_status:
                                machine.current_status = default_status
                                save_fields.append('current_status')
                        updated += 1
                    if save_fields:
                        machine.save(update_fields=save_fields)
                except Exception as e:
                    errors.append({'row': row_num, 'error': str(e)})

        except Exception as e:
            return Response({'detail': f'Ошибка чтения файла: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created': created,
            'updated': updated,
            'errors': errors,
            'total_processed': created + updated + len(errors)
        })

    @extend_schema(summary='Шаблон Excel для импорта')
    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Шаблон импорта'
        headers = ['Инв. №*', 'Наименование*', 'Модель', 'Производитель', 'Год выпуска']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="machines_import_template.xlsx"'
        return response

    @extend_schema(summary='Получить/создать/обновить график ТО станка')
    @action(detail=True, methods=['get', 'post', 'patch', 'delete'], url_path='maintenance')
    def maintenance(self, request, pk=None):
        if request.method == 'DELETE':
            if not request.user.role == 'admin':
                return Response({'detail': 'Только администратор может удалить график ТО'},
                                status=status.HTTP_403_FORBIDDEN)
            try:
                schedule = MaintenanceSchedule.objects.get(machine_id=pk)
                schedule.delete()
                return Response({'detail': 'График ТО удалён'})
            except MaintenanceSchedule.DoesNotExist:
                return Response({'detail': 'График ТО не найден'}, status=status.HTTP_404_NOT_FOUND)

        machine = self.get_object()

        if request.method == 'GET':
            try:
                schedule = MaintenanceSchedule.objects.prefetch_related('tasks').get(machine=machine)
                return Response(MaintenanceScheduleSerializer(schedule).data)
            except MaintenanceSchedule.DoesNotExist:
                return Response(None)

        if not request.user.role == 'admin':
            return Response({'detail': 'Только администратор может задавать график ТО'},
                            status=status.HTTP_403_FORBIDDEN)

        serializer = MaintenanceScheduleWriteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        if request.method == 'POST':
            if hasattr(machine, 'maintenance_schedule'):
                return Response({'detail': 'График ТО уже существует. Используйте PATCH для обновления.'},
                                status=status.HTTP_400_BAD_REQUEST)
            schedule = MaintenanceSchedule.objects.create(
                machine=machine,
                created_by=request.user,
                updated_by=request.user,
                **serializer.validated_data,
            )
        else:
            try:
                schedule = machine.maintenance_schedule
            except MaintenanceSchedule.DoesNotExist:
                return Response({'detail': 'График ТО не найден'}, status=status.HTTP_404_NOT_FOUND)
            for attr, value in serializer.validated_data.items():
                setattr(schedule, attr, value)
            schedule.updated_by = request.user
            schedule.save()

        return Response(
            MaintenanceScheduleSerializer(schedule).data,
            status=status.HTTP_201_CREATED if request.method == 'POST' else status.HTTP_200_OK,
        )

    @extend_schema(summary='История ТО станка')
    @action(detail=True, methods=['get'], url_path='maintenance-history')
    def maintenance_history(self, request, pk=None):
        machine = self.get_object()
        history = MaintenanceHistory.objects.filter(machine=machine).select_related('completed_by')
        return Response(MaintenanceHistorySerializer(history, many=True).data)

    @extend_schema(summary='Начать ремонт станка')
    @action(detail=True, methods=['post'], url_path='maintenance/start-repair')
    def maintenance_start_repair(self, request, pk=None):
        if not request.user.role == 'admin':
            return Response({'detail': 'Только администратор может начать ремонт'},
                            status=status.HTTP_403_FORBIDDEN)
        machine = self.get_object()
        try:
            schedule = machine.maintenance_schedule
        except MaintenanceSchedule.DoesNotExist:
            return Response({'detail': 'График ТО не найден'}, status=status.HTTP_404_NOT_FOUND)
        from django.utils import timezone
        schedule.repair_started_at = timezone.now()
        schedule.save(update_fields=['repair_started_at'])
        return Response(MaintenanceScheduleSerializer(schedule).data)

    @extend_schema(summary='Список/создание задач ремонта')
    @action(detail=True, methods=['get', 'post'], url_path='maintenance/tasks')
    def maintenance_tasks(self, request, pk=None):
        machine = self.get_object()
        try:
            schedule = machine.maintenance_schedule
        except MaintenanceSchedule.DoesNotExist:
            return Response({'detail': 'График ТО не найден'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            tasks = schedule.tasks.select_related('assignee').prefetch_related(
                'spare_parts_used__spare_part__unit'
            ).all()
            # Non-admin: show only tasks assigned to them (now direct User FK)
            if request.user.role != 'admin':
                tasks = tasks.filter(assignee=request.user)
            return Response(RepairTaskSerializer(tasks, many=True, context={'request': request}).data)

        if not request.user.role == 'admin':
            return Response({'detail': 'Faqat administrator vazifa qo\'sha oladi'},
                            status=status.HTTP_403_FORBIDDEN)
        serializer = RepairTaskSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        task = RepairTask.objects.create(
            schedule=schedule,
            title=serializer.validated_data.get('title', ''),
            assignee=serializer.validated_data.get('assignee'),
            due_date=serializer.validated_data.get('due_date'),
        )
        return Response(RepairTaskSerializer(task, context={'request': request}).data,
                        status=status.HTTP_201_CREATED)

    @extend_schema(summary='Обновить/удалить задачу ремонта')
    @action(detail=True, methods=['patch', 'delete'],
            url_path=r'maintenance/tasks/(?P<task_pk>[^/.]+)')
    def maintenance_task_detail(self, request, pk=None, task_pk=None):
        machine = self.get_object()
        try:
            task = RepairTask.objects.get(pk=task_pk, schedule__machine=machine)
        except RepairTask.DoesNotExist:
            return Response({'detail': 'Задача не найдена'}, status=status.HTTP_404_NOT_FOUND)

        is_admin = request.user.role == 'admin'

        if request.method == 'DELETE':
            if not is_admin and task.assignee_id != request.user.id:
                return Response({'detail': 'Faqat administrator yoki biriktirilgan foydalanuvchi o\'chira oladi'},
                                status=status.HTTP_403_FORBIDDEN)
            task.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        # For PATCH: admin always allowed; non-admin only for is_done toggle on their own task
        is_done_only = list(request.data.keys()) == ['is_done']
        if not is_admin:
            if not is_done_only:
                return Response({'detail': 'Faqat administrator vazifani o\'zgartira oladi'},
                                status=status.HTTP_403_FORBIDDEN)
            # Check assignee matches current user (direct User FK comparison)
            if task.assignee_id != request.user.id:
                return Response(
                    {'detail': 'Bu vazifa sizga biriktirilmagan'},
                    status=status.HTTP_403_FORBIDDEN
                )

        from django.utils import timezone
        from django.db import transaction
        from collections import defaultdict
        from decimal import Decimal
        from apps.warehouse.models import SparePart
        from apps.machines.models import TaskSparePart as TSP
        import logging
        logger = logging.getLogger(__name__)

        is_done = request.data.get('is_done')
        if is_done is not None:
            was_done = task.is_done
            marking_done = bool(is_done) and not was_done
            insufficient_error = None

            with transaction.atomic():
                task.is_done = bool(is_done)
                task.done_at = timezone.now() if task.is_done else None
                task.save(update_fields=['is_done', 'done_at'])

                if marking_done:
                    usages = list(
                        task.spare_parts_used.filter(deducted=False)
                        .values('id', 'spare_part_id', 'quantity_used')
                    )
                    totals = defaultdict(Decimal)
                    usages_by_sp = defaultdict(list)
                    for u in usages:
                        totals[u['spare_part_id']] += u['quantity_used']
                        usages_by_sp[u['spare_part_id']].append((u['id'], u['quantity_used']))

                    # PASS 1: Lock + validate (select_for_update works inside atomic)
                    insufficient = []
                    sp_cache = {}
                    for sp_id, total_qty in totals.items():
                        sp = SparePart.objects.select_for_update().get(pk=sp_id)
                        sp_cache[sp_id] = sp
                        if sp.quantity is not None and sp.quantity < total_qty:
                            insufficient.append(
                                f'Omborda "{sp.name}" yetarli emas: '
                                f'kerak {total_qty}, mavjud {sp.quantity}'
                            )

                    if insufficient:
                        insufficient_error = ' | '.join(insufficient)
                        transaction.set_rollback(True)
                    else:
                        # PASS 2: Deduct — only when all validations passed
                        for sp_id, total_qty in totals.items():
                            sp = sp_cache[sp_id]
                            fixed_unit_price = sp.unit_price
                            logger.info('[DEDUCT] %s: before=%s, deduct=%s, unit_price=%s',
                                        sp.name, sp.quantity, total_qty, fixed_unit_price)
                            if sp.quantity is not None:
                                sp.quantity -= total_qty
                                sp.save(update_fields=['quantity', 'updated_at'])
                                logger.info('[DEDUCT] %s: after=%s', sp.name, sp.quantity)
                            for uid, uqty in usages_by_sp[sp_id]:
                                cost = (uqty * fixed_unit_price) if fixed_unit_price is not None else None
                                TSP.objects.filter(pk=uid).update(deducted=True, cost=cost)

                elif not task.is_done and was_done:
                    restore_usages = list(
                        task.spare_parts_used.filter(deducted=True)
                        .values('id', 'spare_part_id', 'quantity_used')
                    )
                    restore_totals = defaultdict(Decimal)
                    restore_ids = defaultdict(list)
                    for u in restore_usages:
                        restore_totals[u['spare_part_id']] += u['quantity_used']
                        restore_ids[u['spare_part_id']].append(u['id'])
                    for sp_id, total_qty in restore_totals.items():
                        sp = SparePart.objects.select_for_update().get(pk=sp_id)
                        if sp.quantity is not None:
                            sp.quantity += total_qty
                            sp.save(update_fields=['quantity', 'updated_at'])
                        TSP.objects.filter(id__in=restore_ids[sp_id]).update(deducted=False)

            if insufficient_error:
                return Response({'detail': insufficient_error}, status=status.HTTP_400_BAD_REQUEST)

        if 'has_bonus' in request.data or 'bonus_amount' in request.data:
            if 'has_bonus' in request.data:
                task.has_bonus = bool(request.data.get('has_bonus'))
            if 'bonus_amount' in request.data:
                bonus_amount = request.data.get('bonus_amount')
                task.bonus_amount = bonus_amount if bonus_amount not in (None, '') else None
            if not task.has_bonus:
                task.bonus_amount = None
            task.save(update_fields=['has_bonus', 'bonus_amount'])

        task.refresh_from_db()
        return Response(RepairTaskSerializer(task, context={'request': request}).data)

    @extend_schema(summary='Список/добавление ehtiyot qism для задачи')
    @action(detail=True, methods=['get', 'post'],
            url_path=r'maintenance/tasks/(?P<task_pk>[^/.]+)/spare-parts')
    def maintenance_task_spare_parts(self, request, pk=None, task_pk=None):
        machine = self.get_object()
        try:
            task = RepairTask.objects.get(pk=task_pk, schedule__machine=machine)
        except RepairTask.DoesNotExist:
            return Response({'detail': 'Vazifa topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'GET':
            usages = task.spare_parts_used.select_related('spare_part__unit').all()
            return Response(TaskSparePartSerializer(usages, many=True).data)

        # Allow admin OR the task assignee
        is_admin = request.user.role == 'admin'
        is_assignee = task.assignee_id == request.user.id
        if not is_admin and not is_assignee:
            return Response(
                {'detail': 'Faqat administrator yoki vazifaga biriktirilgan foydalanuvchi qo\'sha oladi'},
                status=status.HTTP_403_FORBIDDEN
            )

        spare_part_id = request.data.get('spare_part')
        quantity_used = request.data.get('quantity_used')
        if not spare_part_id or not quantity_used:
            return Response({'detail': 'spare_part va quantity_used majburiy'}, status=status.HTTP_400_BAD_REQUEST)

        from apps.warehouse.models import SparePart
        try:
            sp = SparePart.objects.select_related('unit').get(pk=spare_part_id)
        except SparePart.DoesNotExist:
            return Response({'detail': 'Ehtiyot qism topilmadi'}, status=status.HTTP_404_NOT_FOUND)

        from decimal import Decimal, InvalidOperation
        try:
            qty = Decimal(str(quantity_used))
        except InvalidOperation:
            return Response({'detail': 'Noto\'g\'ri miqdor'}, status=status.HTTP_400_BAD_REQUEST)

        if sp.quantity is not None and qty > sp.quantity:
            return Response(
                {'detail': f'Yetarli miqdor yo\'q. Mavjud: {sp.quantity}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        notes = request.data.get('notes', '')
        usage = TaskSparePart.objects.create(task=task, spare_part=sp, quantity_used=qty, notes=notes)
        return Response(TaskSparePartSerializer(usage).data, status=status.HTTP_201_CREATED)

    @extend_schema(summary='Удалить ehtiyot qism из задачи')
    @action(detail=True, methods=['delete'],
            url_path=r'maintenance/tasks/(?P<task_pk>[^/.]+)/spare-parts/(?P<usage_pk>[^/.]+)')
    def maintenance_task_spare_part_delete(self, request, pk=None, task_pk=None, usage_pk=None):
        machine = self.get_object()
        try:
            usage = TaskSparePart.objects.get(pk=usage_pk, task__schedule__machine=machine)
        except TaskSparePart.DoesNotExist:
            return Response({'detail': 'Topilmadi'}, status=status.HTTP_404_NOT_FOUND)
        is_admin = request.user.role == 'admin'
        is_assignee = usage.task.assignee_id == request.user.id
        if not is_admin and not is_assignee:
            return Response({'detail': 'Nет доступа'}, status=status.HTTP_403_FORBIDDEN)
        usage.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @extend_schema(summary='Отметить ТО как выполненное', request=MaintenanceCompleteSerializer)
    @action(detail=True, methods=['post'], url_path='maintenance/complete')
    def maintenance_complete(self, request, pk=None):
        if not request.user.role == 'admin':
            return Response({'detail': 'Только администратор может отмечать выполнение ТО'},
                            status=status.HTTP_403_FORBIDDEN)

        machine = self.get_object()
        try:
            schedule = machine.maintenance_schedule
        except MaintenanceSchedule.DoesNotExist:
            return Response({'detail': 'График ТО не найден'}, status=status.HTTP_404_NOT_FOUND)

        serializer = MaintenanceCompleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # ── Pre-check: validate all undone-task spare parts before completing ──
        from collections import defaultdict
        from decimal import Decimal as D
        from apps.warehouse.models import SparePart as SP

        pending_totals = defaultdict(D)
        for task in schedule.tasks.prefetch_related('spare_parts_used').all():
            if not task.is_done:
                for u in task.spare_parts_used.filter(deducted=False).values('spare_part_id', 'quantity_used'):
                    pending_totals[u['spare_part_id']] += u['quantity_used']

        insufficient = []
        for sp_id, qty in pending_totals.items():
            sp = SP.objects.get(pk=sp_id)
            if sp.quantity is not None and sp.quantity < qty:
                insufficient.append(
                    f'Omborda "{sp.name}" yetarli emas: kerak {qty}, mavjud {sp.quantity}'
                )
        if insufficient:
            return Response(
                {'detail': ' | '.join(insufficient)},
                status=status.HTTP_400_BAD_REQUEST
            )

        from dateutil.relativedelta import relativedelta

        # Build tasks snapshot + calculate total_cost
        from decimal import Decimal as D
        tasks_snapshot = []
        total_cost = D('0')
        for task in schedule.tasks.prefetch_related('spare_parts_used__spare_part__unit').all():
            spare_parts = []
            task_cost = D('0')
            for usage in task.spare_parts_used.all():
                sp = usage.spare_part
                unit_str = sp.unit.short_name or sp.unit.name if sp.unit else ''
                cost = usage.cost or D('0')
                task_cost += cost
                spare_parts.append({
                    'name': sp.name,
                    'quantity_used': str(usage.quantity_used),
                    'unit': unit_str,
                    'unit_price': str(sp.unit_price) if sp.unit_price is not None else None,
                    'cost': str(cost),
                })
            bonus_amount = task.bonus_amount or D('0') if task.has_bonus else D('0')
            task_cost += bonus_amount
            total_cost += task_cost
            tasks_snapshot.append({
                'title': task.title,
                'assignee_name': task.assignee.get_full_name() if task.assignee else None,
                'due_date': task.due_date.isoformat() if task.due_date else None,
                'is_done': task.is_done,
                'spare_parts': spare_parts,
                'has_bonus': task.has_bonus,
                'bonus_amount': str(bonus_amount) if task.has_bonus else None,
                'task_cost': str(task_cost),
            })

        # Create history record
        completion_note = serializer.validated_data.get('notes', '')
        MaintenanceHistory.objects.create(
            machine=machine,
            completed_by=request.user,
            interval_months=schedule.interval_months,
            notes=completion_note,
            tasks_snapshot=tasks_snapshot,
            repair_started_at=schedule.repair_started_at,
            total_cost=total_cost,
        )

        completion_date = serializer.validated_data['completion_date']
        schedule.last_maintenance_date = completion_date
        schedule.next_maintenance_date = completion_date + relativedelta(months=schedule.interval_months)
        if completion_note:
            schedule.notes = completion_note
        schedule.repair_started_at = None
        schedule.updated_by = request.user
        schedule.save()

        return Response(MaintenanceScheduleSerializer(schedule).data)


@extend_schema(tags=['maintenance'])
class MaintenanceAlertsView(generics.ListAPIView):
    """
    Возвращает все графики ТО.
    Без параметров: только станки с истекающим/просроченным ТО (в течение 30 дней).
    С параметром ?all=true: все графики ТО.
    """
    serializer_class = MaintenanceScheduleSerializer

    def get_queryset(self):
        from django.utils import timezone
        import datetime
        qs = MaintenanceSchedule.objects.filter(
            machine__deleted_at__isnull=True,
        ).select_related(
            'machine', 'machine__workshop', 'created_by', 'updated_by'
        ).prefetch_related('tasks').order_by('next_maintenance_date')

        if self.request.query_params.get('all') != 'true':
            today = timezone.now().date()
            warning_threshold = today + datetime.timedelta(days=30)
            qs = qs.filter(next_maintenance_date__lte=warning_threshold)

        return qs

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        data = MaintenanceScheduleSerializer(qs, many=True).data
        overdue = [x for x in data if x['alert_level'] == 'overdue']
        critical = [x for x in data if x['alert_level'] == 'critical']
        warning = [x for x in data if x['alert_level'] == 'warning']
        return Response({
            'total': len(data),
            'overdue_count': len(overdue),
            'critical_count': len(critical),
            'warning_count': len(warning),
            'results': data,
        })


@extend_schema(tags=['maintenance'])
class MaintenanceExportView(generics.GenericAPIView):
    """Экспорт данных ТО в Excel или PDF с фильтрацией."""
    permission_classes = [IsAdmin]

    def get(self, request):
        export_format = request.query_params.get('export_format', 'excel')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        workshop_id = request.query_params.get('workshop_id')
        status_id = request.query_params.get('status_id')

        qs = MaintenanceHistory.objects.select_related(
            'machine', 'machine__workshop', 'machine__current_status', 'completed_by'
        ).order_by('-completed_at')

        if date_from:
            from django.utils.dateparse import parse_date
            import datetime as dt
            d = parse_date(date_from)
            if d:
                from django.utils.timezone import make_aware
                qs = qs.filter(completed_at__gte=make_aware(dt.datetime.combine(d, dt.time.min)))

        if date_to:
            from django.utils.dateparse import parse_date
            import datetime as dt
            d = parse_date(date_to)
            if d:
                from django.utils.timezone import make_aware
                qs = qs.filter(completed_at__lte=make_aware(dt.datetime.combine(d, dt.time.max)))

        if workshop_id:
            qs = qs.filter(machine__workshop_id=workshop_id)

        if status_id:
            qs = qs.filter(machine__current_status_id=status_id)

        records = list(qs)

        if export_format == 'pdf':
            return self._export_pdf(records)
        return self._export_excel(records)

    # ── helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _thin_border():
        s = Side(style='thin', color='D1D5DB')
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _build_workshop_groups(records):
        groups = defaultdict(lambda: defaultdict(lambda: {'name': '', 'inv': '', 'count': 0, 'cost': Decimal('0')}))
        for rec in records:
            ws_name = rec.machine.workshop.name if rec.machine.workshop else 'Без цеха'
            mid = rec.machine.id
            groups[ws_name][mid]['name'] = rec.machine.name
            groups[ws_name][mid]['inv'] = rec.machine.inventory_number
            groups[ws_name][mid]['count'] += 1
            groups[ws_name][mid]['cost'] += rec.total_cost
        return groups

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _export_excel(self, records):
        wb = openpyxl.Workbook()
        border = self._thin_border()

        hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        hdr_fill  = PatternFill(fill_type='solid', fgColor='3B4FC8')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(vertical='center', wrap_text=True)
        alt_fill   = PatternFill(fill_type='solid', fgColor='F0F4FF')
        norm_font  = Font(name='Arial', size=9)
        bold_font  = Font(name='Arial', bold=True, size=9)

        # ── Sheet 1: История ТО ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'История ТО'

        headers1 = ['#', 'Станок', 'Инв. №', 'Цех', 'Выполнил',
                    'Начало ремонта', 'Дата завершения', 'Инт. (мес.)', 'Задания', 'Итого, USD']
        widths1   = [5, 28, 14, 20, 22, 18, 18, 12, 12, 14]

        ws1.row_dimensions[1].height = 36
        for c, (h, w) in enumerate(zip(headers1, widths1), 1):
            cell = ws1.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border
            ws1.column_dimensions[get_column_letter(c)].width = w

        for row_i, rec in enumerate(records, 2):
            tasks = rec.tasks_snapshot or []
            done  = sum(1 for t in tasks if t.get('is_done'))
            row_fill = alt_fill if row_i % 2 == 0 else None

            row_data = [
                row_i - 1,
                rec.machine.name,
                rec.machine.inventory_number,
                rec.machine.workshop.name if rec.machine.workshop else '—',
                rec.completed_by.get_full_name() if rec.completed_by else '—',
                rec.repair_started_at.strftime('%d.%m.%Y %H:%M') if rec.repair_started_at else '—',
                rec.completed_at.strftime('%d.%m.%Y'),
                rec.interval_months,
                f'{done}/{len(tasks)}' if tasks else '—',
                float(rec.total_cost),
            ]
            ws1.row_dimensions[row_i].height = 20
            for c, val in enumerate(row_data, 1):
                cell = ws1.cell(row=row_i, column=c, value=val)
                cell.font = norm_font
                cell.border = border
                cell.alignment = cell_align
                if row_fill:
                    cell.fill = row_fill
                if c == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if c == 10:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')

        # ── Sheet 2: Расходы ─────────────────────────────────────────────────
        ws2 = wb.create_sheet('Расходы')

        headers2 = ['Цех / Станок', 'Инв. №', 'Кол-во ТО', 'Расходы, USD']
        widths2   = [42, 16, 14, 18]

        ws2.row_dimensions[1].height = 36
        for c, (h, w) in enumerate(zip(headers2, widths2), 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border
            ws2.column_dimensions[get_column_letter(c)].width = w

        ws_hdr_fill = PatternFill(fill_type='solid', fgColor='E8EAF6')
        ws_hdr_font = Font(name='Arial', bold=True, size=9, color='1A237E')
        sub_fill    = PatternFill(fill_type='solid', fgColor='EDE7F6')
        sub_font    = Font(name='Arial', bold=True, size=9, color='4A148C')
        grand_fill  = PatternFill(fill_type='solid', fgColor='3B4FC8')
        grand_font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')

        cur_row = 2
        grand_cnt  = 0
        grand_cost = Decimal('0')
        groups = self._build_workshop_groups(records)

        for ws_name in sorted(groups):
            # workshop header
            ws2.row_dimensions[cur_row].height = 22
            row_data = [ws_name, '', '', '']
            for c, val in enumerate(row_data, 1):
                cell = ws2.cell(row=cur_row, column=c, value=val)
                cell.font = ws_hdr_font
                cell.fill = ws_hdr_fill
                cell.border = border
                cell.alignment = Alignment(vertical='center')
            cur_row += 1

            ws_cnt  = 0
            ws_cost = Decimal('0')
            for mid, d in sorted(groups[ws_name].items(), key=lambda x: x[1]['name']):
                ws2.row_dimensions[cur_row].height = 18
                row_data = ['    ' + d['name'], d['inv'], d['count'], float(d['cost'])]
                for c, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=cur_row, column=c, value=val)
                    cell.font = norm_font
                    cell.border = border
                    cell.alignment = cell_align
                    if c == 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    if c == 4:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                ws_cnt  += d['count']
                ws_cost += d['cost']
                cur_row += 1

            # workshop subtotal
            ws2.row_dimensions[cur_row].height = 22
            row_data = [f'Итого: {ws_name}', '', ws_cnt, float(ws_cost)]
            for c, val in enumerate(row_data, 1):
                cell = ws2.cell(row=cur_row, column=c, value=val)
                cell.font = sub_font
                cell.fill = sub_fill
                cell.border = border
                cell.alignment = cell_align
                if c == 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if c == 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            grand_cnt  += ws_cnt
            grand_cost += ws_cost
            cur_row += 1

        # grand total
        ws2.row_dimensions[cur_row].height = 26
        for c, val in enumerate(['ИТОГО:', '', grand_cnt, float(grand_cost)], 1):
            cell = ws2.cell(row=cur_row, column=c, value=val)
            cell.font = grand_font
            cell.fill = grand_fill
            cell.border = border
            cell.alignment = cell_align
            if c == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="TO_export_{ts}.xlsx"'
        return response

    # ── PDF ──────────────────────────────────────────────────────────────────

    def _export_pdf(self, records):
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # Register Cyrillic-capable font
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
        font_candidates = [
            ('C:/Windows/Fonts/arial.ttf',        'C:/Windows/Fonts/arialbd.ttf'),
            ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
             '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
            ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
             '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ]
        for reg_path, bold_path in font_candidates:
            if os.path.exists(reg_path):
                try:
                    pdfmetrics.registerFont(TTFont('ExportFont', reg_path))
                    font_name = 'ExportFont'
                except Exception:
                    pass
                if os.path.exists(bold_path):
                    try:
                        pdfmetrics.registerFont(TTFont('ExportFontBold', bold_path))
                        font_bold = 'ExportFontBold'
                    except Exception:
                        pass
                break

        buf = io.BytesIO()
        page_w, page_h = landscape(A4)
        avail_w = page_w - 30 * mm

        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=20*mm,
        )

        def style(name, **kw):
            kw.setdefault('fontName', font_name)
            return ParagraphStyle(name, **kw)

        title_s  = style('t', fontName=font_bold, fontSize=14, textColor=rl_colors.HexColor('#1e3a5f'), spaceAfter=4)
        sub_s    = style('s', fontSize=9, textColor=rl_colors.HexColor('#64748b'), spaceAfter=10)
        sec_s    = style('sec', fontName=font_bold, fontSize=11, textColor=rl_colors.HexColor('#1e3a5f'), spaceBefore=10, spaceAfter=6)
        sig_s    = style('sig', fontSize=9, textColor=rl_colors.HexColor('#374151'), spaceAfter=6)

        now = datetime.now()
        story = [
            Paragraph('Lazana — Техническое обслуживание', title_s),
            Paragraph(f'Дата формирования: {now.strftime("%d.%m.%Y %H:%M")}', sub_s),
            HRFlowable(width='100%', thickness=1, color=rl_colors.HexColor('#e2e8f0'), spaceAfter=10),
        ]

        th_bg = rl_colors.HexColor('#3B4FC8')
        th_fg = rl_colors.white
        alt_bg = rl_colors.HexColor('#F0F4FF')
        grid_c = rl_colors.HexColor('#D1D5DB')

        base_table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), th_bg),
            ('TEXTCOLOR',  (0, 0), (-1, 0), th_fg),
            ('FONTNAME',   (0, 0), (-1, 0), font_bold),
            ('FONTSIZE',   (0, 0), (-1, 0), 8),
            ('FONTNAME',   (0, 1), (-1, -1), font_name),
            ('FONTSIZE',   (0, 1), (-1, -1), 7.5),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',       (0, 0), (-1, -1), 0.5, grid_c),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ]

        # ── Table 1: История ТО ──────────────────────────────────────────────
        story.append(Paragraph('История ТО', sec_s))

        t1_headers = ['#', 'Станок', 'Инв. №', 'Цех', 'Выполнил', 'Завершён', 'Инт.', 'Задания', 'USD']
        t1_widths  = [w * avail_w for w in [0.04, 0.19, 0.10, 0.12, 0.16, 0.10, 0.06, 0.09, 0.10]]

        t1_data = [t1_headers]
        t1_style_cmds = list(base_table_style) + [
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (6, 0), (8, -1), 'CENTER'),
            ('ALIGN', (8, 0), (8, -1), 'RIGHT'),
        ]

        for i, rec in enumerate(records, 1):
            tasks = rec.tasks_snapshot or []
            done  = sum(1 for t in tasks if t.get('is_done'))
            if i % 2 == 0:
                t1_style_cmds.append(('BACKGROUND', (0, i), (-1, i), alt_bg))
            t1_data.append([
                str(i),
                rec.machine.name[:35],
                rec.machine.inventory_number,
                (rec.machine.workshop.name[:18] if rec.machine.workshop else '—'),
                (rec.completed_by.get_full_name()[:22] if rec.completed_by else '—'),
                rec.completed_at.strftime('%d.%m.%Y'),
                str(rec.interval_months),
                f'{done}/{len(tasks)}' if tasks else '—',
                f'{float(rec.total_cost):.2f}',
            ])

        t1 = Table(t1_data, colWidths=t1_widths, repeatRows=1)
        t1.setStyle(TableStyle(t1_style_cmds))
        story.append(t1)
        story.append(Spacer(1, 8*mm))

        # ── Table 2: Расходы ─────────────────────────────────────────────────
        story.append(Paragraph('Расходы по цехам', sec_s))

        t2_headers = ['Цех / Станок', 'Инв. №', 'Кол-во ТО', 'Расходы, USD']
        t2_widths  = [w * avail_w for w in [0.50, 0.16, 0.15, 0.19]]

        ws_hdr_bg  = rl_colors.HexColor('#E8EAF6')
        sub_bg     = rl_colors.HexColor('#EDE7F6')
        grand_bg   = rl_colors.HexColor('#3B4FC8')

        t2_data = [t2_headers]
        t2_style_cmds = list(base_table_style) + [
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
        ]

        groups = self._build_workshop_groups(records)
        grand_cnt  = 0
        grand_cost = Decimal('0')
        ws_hdr_rows = []
        sub_rows    = []

        for ws_name in sorted(groups):
            t2_data.append([ws_name, '', '', ''])
            ws_hdr_rows.append(len(t2_data) - 1)

            ws_cnt  = 0
            ws_cost = Decimal('0')
            for mid, d in sorted(groups[ws_name].items(), key=lambda x: x[1]['name']):
                t2_data.append([f'   {d["name"]}', d['inv'], str(d['count']), f'{float(d["cost"]):.2f}'])
                ws_cnt  += d['count']
                ws_cost += d['cost']

            t2_data.append([f'Итого: {ws_name}', '', str(ws_cnt), f'{float(ws_cost):.2f}'])
            sub_rows.append(len(t2_data) - 1)
            grand_cnt  += ws_cnt
            grand_cost += ws_cost

        t2_data.append(['ИТОГО:', '', str(grand_cnt), f'{float(grand_cost):.2f}'])
        grand_row = len(t2_data) - 1

        for r in ws_hdr_rows:
            t2_style_cmds += [
                ('BACKGROUND', (0, r), (-1, r), ws_hdr_bg),
                ('FONTNAME',   (0, r), (-1, r), font_bold),
                ('TEXTCOLOR',  (0, r), (-1, r), rl_colors.HexColor('#1A237E')),
            ]
        for r in sub_rows:
            t2_style_cmds += [
                ('BACKGROUND', (0, r), (-1, r), sub_bg),
                ('FONTNAME',   (0, r), (-1, r), font_bold),
                ('TEXTCOLOR',  (0, r), (-1, r), rl_colors.HexColor('#4A148C')),
            ]
        t2_style_cmds += [
            ('BACKGROUND', (0, grand_row), (-1, grand_row), grand_bg),
            ('TEXTCOLOR',  (0, grand_row), (-1, grand_row), rl_colors.white),
            ('FONTNAME',   (0, grand_row), (-1, grand_row), font_bold),
            ('FONTSIZE',   (0, grand_row), (-1, grand_row), 9),
        ]

        t2 = Table(t2_data, colWidths=t2_widths, repeatRows=1)
        t2.setStyle(TableStyle(t2_style_cmds))
        story.append(t2)

        # ── Signature ────────────────────────────────────────────────────────
        story.append(Spacer(1, 14*mm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=rl_colors.HexColor('#CBD5E1'), spaceAfter=8))
        story.append(Paragraph('Ответственный: _________________________ / _________________________', sig_s))
        story.append(Paragraph('Дата: _________________________', sig_s))

        doc.build(story)
        buf.seek(0)
        ts = now.strftime('%Y%m%d_%H%M')
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="TO_export_{ts}.pdf"'
        return resp
